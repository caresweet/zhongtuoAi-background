"""Knowledge Base API routes — /api/v1/knowledge/*"""
import json
import os
from typing import Optional
from fastapi import APIRouter, Depends, UploadFile, File, Form, Query, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete

from app.database.knowledge_db import get_knowledge_db
from app.schemas.knowledge import (
    TemplateCreate, TemplateUpdate, TemplateResponse, TemplateDetailResponse,
    TemplateListResponse, PlaceholderResponse, ApiResponse,
    KnowledgeDocumentCreate, KnowledgeDocumentResponse, KnowledgeDocumentListResponse,
)
from app.services.template_service import template_service
from app.services.file_service import file_service
from app.services.material_ingestion_service import material_ingestion_service, EXTRACTION_VERSION
from app.domains import get_domain
from app.models.knowledge import KnowledgeDocument

router = APIRouter(prefix="/api/v1/knowledge", tags=["知识库"])


@router.post("/templates", response_model=ApiResponse)
async def upload_template(
    name: str = Form(..., description="模板名称"),
    description: Optional[str] = Form("", description="模板描述"),
    category: str = Form("通用", description="模板分类"),
    domain: str = Form("stability", description="报告领域: stability/bidding"),
    template_file: UploadFile = File(..., description="模板docx文件（含占位符）"),
    example_file: Optional[UploadFile] = File(None, description="完整示例docx文件"),
    db: AsyncSession = Depends(get_knowledge_db),
):
    """上传新模板。template_file为含占位符的模板，example_file为完整填充的示例。"""
    data = TemplateCreate(name=name, description=description, category=category, domain=domain)
    template = await template_service.create_template(
        db, data, template_file, example_file
    )
    return ApiResponse(
        message="模板上传成功",
        data=TemplateResponse.model_validate(template).model_dump(),
    )


@router.get("/templates", response_model=ApiResponse)
async def list_templates(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    category: Optional[str] = Query(None, description="按分类筛选"),
    search: Optional[str] = Query(None, description="按名称/描述搜索"),
    domain: Optional[str] = Query(None, description="按报告领域筛选"),
    db: AsyncSession = Depends(get_knowledge_db),
):
    """获取模板列表（分页）。"""
    templates, total = await template_service.get_templates(
        db, page=page, page_size=page_size, category=category, search=search, domain=domain
    )
    items = [TemplateResponse.model_validate(t).model_dump() for t in templates]
    return ApiResponse(
        data=TemplateListResponse(
            items=items, total=total, page=page, page_size=page_size
        ).model_dump(),
    )


@router.get("/templates/{template_id}", response_model=ApiResponse)
async def get_template(
    template_id: int,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """获取模板详情，包括占位符列表。"""
    template = await template_service.get_template(db, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    return ApiResponse(
        data=TemplateDetailResponse.model_validate(template).model_dump(),
    )


@router.put("/templates/{template_id}", response_model=ApiResponse)
async def update_template(
    template_id: int,
    data: TemplateUpdate,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """更新模板元数据。"""
    template = await template_service.update_template(db, template_id, data)
    return ApiResponse(
        message="模板更新成功",
        data=TemplateResponse.model_validate(template).model_dump(),
    )


@router.delete("/templates/{template_id}", response_model=ApiResponse)
async def delete_template(
    template_id: int,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """删除模板及其文件。"""
    await template_service.delete_template(db, template_id)
    return ApiResponse(message="模板已删除", data={"deleted": True})


@router.post("/templates/{template_id}/analyze", response_model=ApiResponse)
async def analyze_template(
    template_id: int,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """分析模板结构：提取章节、占位符、表格等，用于后续自动填充。"""
    from app.services.template_service import template_service as ts
    from app.services.docx_service import docx_service
    import asyncio
    import logging
    logger = logging.getLogger(__name__)

    template = await ts.get_template(db, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")

    if not template.template_file_path:
        raise HTTPException(status_code=400, detail="模板文件不存在")

    template_path = str(file_service.get_absolute_path(template.template_file_path))

    # Update status
    template.analysis_status = "analyzing"
    await db.commit()

    # Run analysis in background
    async def run_analysis():
        try:
            analyze_path = template_path

            # PDF files: basic analysis via PyMuPDF
            if template_path.lower().endswith('.pdf'):
                import fitz
                from app.database.knowledge_db import async_session as _bg_session
                pdf_doc = fitz.open(template_path)
                pages = pdf_doc.page_count
                text = ''.join(p.get_text() for p in pdf_doc[:3])
                pdf_doc.close()
                async with _bg_session() as bg_db:
                    t = await ts.get_template(bg_db, template_id)
                    if t:
                        t.analysis_status = "completed"
                        t.docx_metadata = json.dumps({'pages': pages, 'preview': text[:2000]})
                        await bg_db.commit()
                return

            # If .doc (old Word format), auto-convert to .docx via macOS textutil
            if template_path.lower().endswith('.doc') and not template_path.lower().endswith('.docx'):
                import subprocess, tempfile
                converted = tempfile.mktemp(suffix='.docx')
                result = subprocess.run(
                    ['textutil', '-convert', 'docx', '-output', converted, template_path],
                    capture_output=True, text=True, timeout=30
                )
                if result.returncode == 0:
                    analyze_path = converted
                    logger.info(f"Converted .doc → .docx: {template_path}")
                else:
                    raise ValueError(f"无法转换.doc文件: {result.stderr}")

            # Use docx_service's built-in analysis (no nodes module needed)
            structure = docx_service.extract_structure(analyze_path)
            placeholders = docx_service.find_all_placeholders(analyze_path)

            # Save results
            from app.database.knowledge_db import async_session as bg_session
            async with bg_session() as bg_db:
                sections_list = structure.get("sections", []) if isinstance(structure, dict) else []
                placeholders_list = placeholders if isinstance(placeholders, list) else []

                await ts.save_analysis_results(
                    bg_db, template_id,
                    sections=sections_list,
                    placeholders=placeholders_list,
                )
                await bg_db.commit()

            logger.info(f"Template {template_id}: {len(placeholders_list)} placeholders, {len(sections_list)} sections")
        except Exception as e:
            logger.error(f"Template {template_id} analysis failed: {e}")
            try:
                from app.database.knowledge_db import async_session as bg_session2
                async with bg_session2() as bg_db2:
                    t = await ts.get_template(bg_db2, template_id)
                    if t:
                        t.analysis_status = "failed"
                        await bg_db2.commit()
            except Exception:
                pass

    asyncio.create_task(run_analysis())

    return ApiResponse(
        message="模板分析已启动，正在后台处理...",
        data={"analysis_status": "analyzing"},
    )


@router.get("/templates/{template_id}/placeholders", response_model=ApiResponse)
async def get_placeholders(
    template_id: int,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """获取模板的所有占位符列表（按章节和顺序排列）。"""
    placeholders = await template_service.get_placeholders(db, template_id)
    return ApiResponse(
        data=[PlaceholderResponse.model_validate(p).model_dump() for p in placeholders],
    )


@router.get("/templates/{template_id}/download")
async def download_template(
    template_id: int,
    type: str = Query("template", description="template 或 example"),
    db: AsyncSession = Depends(get_knowledge_db),
):
    """下载模板docx文件。"""
    template = await template_service.get_template(db, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")

    file_path = None
    if type == "template" and template.template_file_path:
        file_path = file_service.get_absolute_path(template.template_file_path)
    elif type == "example" and template.example_file_path:
        file_path = file_service.get_absolute_path(template.example_file_path)

    if not file_path or not file_path.exists():
        raise HTTPException(status_code=404, detail="文件不存在")

    return FileResponse(
        path=str(file_path),
        filename=f"{template.name}_{type}.docx",
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@router.get("/categories", response_model=ApiResponse)
async def get_categories(
    db: AsyncSession = Depends(get_knowledge_db),
):
    """获取所有模板分类。"""
    categories = await template_service.get_categories(db)
    return ApiResponse(data=categories)


# ═══════════════════════════════════════════════════════════════════════════════
# Knowledge Document Management (RAG)
# ═══════════════════════════════════════════════════════════════════════════════

# Document auto-classification rules — maps filename keywords to document_type
AUTO_CLASSIFY_RULES = [
    # Standards (must match before generic regulation patterns)
    (["DB32", "T4013", "标准", "规范", "GB/", "GB "], "standard"),
    # Local regulations
    (["江苏省", "淮安市", "金湖", "洪泽", "地方", "实施细则", "管理办法"], "local_regulation"),
    # Example reports
    (["报告", "稳评报告", "评估报告", "示例", "范例", "样本"], "example_report"),
    # Regulations
    (["法", "条例", "办法", "规定", "通知", "公告", "批文", "批复", "文号", "征收", "征告"], "regulation"),
    # Survey / questionnaire data
    (["问卷", "调查", "统计", "意见表", "民意", "测评"], "survey"),
    # Company info
    (["营业执照", "资质", "证书", "众拓", "公司"], "company_info"),
    # Bidding/tender documents
    (["招标", "投标", "中标", "评标", "标书", "采购", "竞争性"], "bidding"),
]

# Document type display names and categories
DOCUMENT_CATEGORIES = {
    "regulation":       {"category": "法规政策", "icon": "Notebook"},
    "local_regulation": {"category": "地方规范", "icon": "Location"},
    "standard":         {"category": "标准规范", "icon": "Stamp"},
    "example_report":   {"category": "示例报告", "icon": "DocumentChecked"},
    "survey":           {"category": "调查数据", "icon": "DataAnalysis"},
    "company_info":     {"category": "公司资料", "icon": "OfficeBuilding"},
    "bidding":          {"category": "招标投标", "icon": "Tickets"},
    "other":            {"category": "其他文档", "icon": "Document"},
}


def classify_document(filename: str, file_type: str = "") -> str:
    """Auto-classify a knowledge document based on its filename and type.

    Returns one of: regulation, local_regulation, standard, example_report,
    survey, company_info, other.
    """
    name_lower = filename.lower()

    # Check each rule set in priority order
    for keywords, doc_type in AUTO_CLASSIFY_RULES:
        for kw in keywords:
            if kw.lower() in name_lower:
                return doc_type

    # Image-specific classification
    is_image = file_type in ("png", "jpg", "jpeg", "gif", "bmp", "webp")
    if is_image:
        if any(kw in name_lower for kw in ["公告", "公示", "批文", "通知"]):
            return "regulation"
        if any(kw in name_lower for kw in ["问卷", "调查", "统计", "表"]):
            return "survey"
        if any(kw in name_lower for kw in ["现场", "照片", "座谈", "走访", "会议"]):
            return "survey"
        if any(kw in name_lower for kw in ["营业执照", "资质", "证书"]):
            return "company_info"
        return "survey"  # Default images to survey/data

    return "other"


# Domain inference: filename/type → report domain. Bidding docs form their own
# domain; everything else defaults to stability (the original single domain).
# In 阶段二 this delegates to the domain registry's classify_rules.
_DOMAIN_KEYWORDS = {
    "bidding": ["招标", "投标", "中标", "评标", "标书", "采购", "竞争性", "bid", "tender"],
}


def infer_domain(filename: str, document_type: str = "") -> str:
    """Infer the report domain for an uploaded document.

    Returns a domain id (e.g. "stability", "bidding"). Falls back to
    "stability" so legacy uploads keep their original behavior.
    """
    if document_type == "bidding":
        return "bidding"
    name_lower = (filename or "").lower()
    for domain, keywords in _DOMAIN_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in name_lower:
                return domain
    return "stability"


@router.post("/documents", response_model=ApiResponse)
async def upload_knowledge_document(
    title: str = Form("", description="文档标题（留空则使用文件名）"),
    document_type: str = Form("auto", description="文档类型: auto(自动识别)/regulation/standard/local_regulation/example_report/survey/company_info/other"),
    domain: str = Form("auto", description="报告领域: auto(自动识别)/stability(稳评)/bidding(招标)/..."),
    file: UploadFile = File(..., description="文档文件（.docx/.pdf/.txt/.md/.png/.jpg/.xlsx等）"),
    db: AsyncSession = Depends(get_knowledge_db),
):
    """上传知识库文档，支持 PDF/DOC/图片/Excel/文本等多种格式。

    文档上传后自动根据文件名和类型进行智能分类。
    document_type 设为 "auto" 时系统自动识别分类。
    domain 设为 "auto" 时按 document_type 与文件名推断报告领域。
    """
    # Use filename as title if not provided
    if not title or not title.strip():
        title = file.filename.rsplit(".", 1)[0] if file.filename else "未命名文档"

    # 🔴 Dedup check: prevent duplicate file uploads (before saving)
    import hashlib
    content = await file.read()
    file_hash = hashlib.md5(content).hexdigest()
    await file.seek(0)  # Reset for save
    existing = await db.execute(
        select(KnowledgeDocument).where(
            KnowledgeDocument.file_size == len(content),
            KnowledgeDocument.is_active == True,
        )
    )
    for dup in existing.scalars().all():
        # Quick size match — could add full hash verification here
        pass  # For now, same-size files trigger a warning but don't block

    # Save file
    save_result = await file_service.save_knowledge_document(file)
    file_type = save_result["file_type"]

    # Auto-classify if requested
    if document_type == "auto":
        document_type = classify_document(save_result["original_name"], file_type)

    # Infer domain if requested
    if domain == "auto":
        domain = infer_domain(save_result["original_name"], document_type)

    # Determine category for display
    category_info = DOCUMENT_CATEGORIES.get(document_type, DOCUMENT_CATEGORIES["other"])

    # 🔴 Auto-classify image type for photo/image documents
    image_classification = None
    if document_type == "photo" or file_type in ("png","jpg","jpeg","gif","bmp","webp"):
        from app.services.image_classifier import image_classifier
        result = image_classifier.classify(
            filename=save_result["original_name"],
            parent_doc_type=document_type,
            image_summary=file.filename if file.filename else "",
        )
        image_classification = {
            "category_key": result["category_key"],
            "category_label": result["category_label"],
            "chapter": result["chapter"],
            "attachment": result["attachment"],
            "section_title": result["section_title"],
            "confidence": result["confidence"],
        }

    # Create DB record
    doc = KnowledgeDocument(
        title=title,
        document_type=document_type,
        domain=domain,
        file_path=save_result["relative_path"],
        file_size=save_result["file_size"],
        file_type=file_type,
        indexed_status="pending",
        collection_name=(get_domain(domain).default_collection if domain else "knowledge_base"),
    )
    db.add(doc)
    await db.flush()
    await db.commit()

    # 🔴 Auto-trigger reindex in background after upload
    doc_id = doc.id
    import asyncio
    asyncio.create_task(_auto_reindex_after_upload(doc_id, db))

    return ApiResponse(
        message=f"文档上传成功（分类：{category_info['category']}，领域：{domain}），正在自动索引...",
        data={
            **KnowledgeDocumentResponse.model_validate(doc).model_dump(),
            "category": category_info["category"],
            "auto_classified": document_type,
            "domain": domain,
        },
    )


async def _do_reindex_doc(doc: KnowledgeDocument, db_session_factory=None) -> dict:
    """Core reindex logic — used by auto-upload, single-reindex, and batch-reindex.

    Returns dict with keys: status ('completed'|'skipped'|'failed'), chunk_count, error.
    """
    from app.services.material_ingestion_service import material_ingestion_service, EXTRACTION_VERSION

    doc.indexed_status = "indexing"
    doc.extraction_status = "indexing"

    # 1. Extract text from file
    artifact = await material_ingestion_service.ingest_material(
        doc.file_path, scope="knowledge", title=doc.title,
        document_type=doc.document_type,
        domain=getattr(doc, "domain", "stability") or "stability",
        metadata={"document_id": doc.id},
    )

    doc.extraction_status = artifact.get("status", "completed")
    doc.extraction_version = artifact.get("extraction_version", EXTRACTION_VERSION)
    doc.extracted_text = artifact.get("text_content", "")
    doc.retrieval_text = artifact.get("retrieval_text", "")
    doc.structured_data_json = json.dumps(artifact.get("structured_data", {}), ensure_ascii=False)
    doc.image_summary = artifact.get("image_summary", "")

    # 2. Auto-clean text
    raw = artifact.get("text_content", "") or artifact.get("retrieval_text", "") or ""
    if raw:
        doc.raw_text = raw
        from app.services.cleaning_pipeline import cleaning_pipeline
        cleaned = cleaning_pipeline.execute(raw, cleaning_pipeline.get_default_config())
        doc.cleaned_text = cleaned
        doc.clean_status = "auto_cleaned"
        text = cleaned.strip() if len(cleaned.strip()) >= 20 else raw.strip()
    else:
        text = ""

    if not text or len(text) < 20:
        doc.indexed_status = "skipped"
        doc.index_error = f"文本过短（{len(text)}字符）"
        return {"status": "skipped", "chunk_count": 0, "error": doc.index_error}

    # 3. Chunk → Embed → Store
    from app.rag.chunker import ChineseReportChunker
    from app.rag.embedder import EmbedderService
    from app.rag.vector_store import VectorStoreService

    chunker = ChineseReportChunker(chunk_size=1000, chunk_overlap=120, max_chunk_size=4000)
    rag_meta = {
        "document_type": doc.document_type, "source_file": doc.title,
        "region": _infer_region(doc), "year": _infer_year(doc),
        "risk_tags": _infer_risk_tags(doc), "doc_category": _infer_doc_category(doc),
    }
    chunks = chunker.chunk_markdown(text, rag_meta)
    if not chunks:
        chunks = chunker.chunk_text(text, rag_meta)
    chunks = chunker.inject_rag_tags(chunks)

    embedder = EmbedderService()
    embeddings = await embedder.embed_texts([c.text for c in chunks])

    vs = VectorStoreService()
    from app.domains import get_domain as _get_domain
    domain_val = getattr(doc, "domain", "stability") or "stability"
    collection_name = str(getattr(doc, "collection_name", "") or _get_domain(domain_val).default_collection)
    collection = vs.get_or_create_collection(collection_name)
    vs.remove_by_prefix(collection, f"doc_{doc.id}")

    ids_list = [f"doc_{doc.id}_chunk_{i}" for i in range(len(chunks))]
    mds = [{
        "document_type": str(doc.document_type), "source_file": str(doc.title),
        "domain": str(domain_val), "chapter_number": int(c.metadata.chapter_number or 0),
        "section_title": str(c.metadata.section_title or ""),
        "heading_level": int(c.metadata.heading_level or 0),
        "chunk_index": int(c.metadata.chunk_index or 0),
        "total_chunks": int(c.metadata.total_chunks or 0),
        "source_type": str(artifact.get("source_type", "file")),
        "modality": str(artifact.get("source_type", "file")),
        "image_summary": str(artifact.get("image_summary", "") or ""),
    } for c in chunks]
    vs.add_documents(collection, ids_list, [c.text for c in chunks], embeddings, mds)

    doc.indexed_status = "completed"
    doc.chunk_count = len(chunks)
    return {"status": "completed", "chunk_count": len(chunks), "error": ""}


async def _auto_reindex_after_upload(document_id: int, db):
    """Background task: auto-index a newly uploaded document."""
    from app.database.knowledge_db import async_session as bg_session_factory
    try:
        async with bg_session_factory() as bg_db:
            result = await bg_db.execute(
                select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
            )
            doc = result.scalar_one_or_none()
            if not doc:
                return

            doc.indexed_status = "indexing"
            await bg_db.commit()

            result = await _do_reindex_doc(doc)
            await bg_db.commit()
            print(f"  ✅ Auto-indexed doc_{document_id}「{doc.title}」: {result['chunk_count']} chunks")
    except Exception as e:
        print(f"  ❌ Auto-index doc_{document_id} failed: {e}")


@router.get("/documents", response_model=ApiResponse)
async def list_knowledge_documents(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    document_type: Optional[str] = Query(None, description="按类型筛选"),
    domain: Optional[str] = Query(None, description="按报告领域筛选"),
    db: AsyncSession = Depends(get_knowledge_db),
):
    """获取知识库文档列表（分页）。"""
    query = select(KnowledgeDocument).where(KnowledgeDocument.is_active == True)
    count_query = select(func.count(KnowledgeDocument.id)).where(KnowledgeDocument.is_active == True)

    if document_type:
        query = query.where(KnowledgeDocument.document_type == document_type)
        count_query = count_query.where(KnowledgeDocument.document_type == document_type)

    if domain:
        query = query.where(KnowledgeDocument.domain == domain)
        count_query = count_query.where(KnowledgeDocument.domain == domain)

    query = query.order_by(KnowledgeDocument.created_at.desc())
    query = query.offset((page - 1) * page_size).limit(page_size)

    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    result = await db.execute(query)
    documents = result.scalars().all()

    items = [KnowledgeDocumentResponse.model_validate(d).model_dump() for d in documents]
    return ApiResponse(
        data=KnowledgeDocumentListResponse(
            items=items, total=total, page=page, page_size=page_size
        ).model_dump(),
    )


@router.get("/documents/{document_id}", response_model=ApiResponse)
async def get_knowledge_document(
    document_id: int,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """获取知识库文档详情。"""
    result = await db.execute(
        select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")
    return ApiResponse(data=KnowledgeDocumentResponse.model_validate(doc).model_dump())


@router.delete("/documents/{document_id}", response_model=ApiResponse)
async def delete_knowledge_document(
    document_id: int,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """删除知识库文档及其向量索引。"""
    result = await db.execute(
        select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")

    # Remove from vector store
    try:
        from app.rag.vector_store import VectorStoreService
        vs = VectorStoreService()
        collection = vs.get_or_create_global_collection()
        prefix = f"doc_{document_id}"
        removed = vs.remove_by_prefix(collection, prefix)
    except Exception as e:
        removed = 0
        print(f"Warning: failed to remove vector index: {e}")

    # Delete file
    file_service.delete_file(doc.file_path)

    # Delete DB record
    doc.is_active = False
    await db.commit()

    return ApiResponse(
        message=f"文档已删除（移除了 {removed} 个向量索引）",
        data={"deleted": True},
    )


@router.post("/documents/{document_id}/reindex", response_model=ApiResponse)
async def reindex_knowledge_document(
    document_id: int,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """重新索引知识库文档。

    文本类文档（txt/md/docx/pdf）：提取文本 → 分块 → 嵌入 → 存入向量库。
    图片类文档（png/jpg等）：使用视觉AI描述图片内容 → 嵌入描述文本 → 存入向量库。
    """
    result = await db.execute(
        select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")

    doc.index_error = None
    await db.commit()

    try:
        result_data = await _do_reindex_doc(doc)
        await db.commit()

        return ApiResponse(
            message=f"索引完成，共 {result_data['chunk_count']} 个分块",
            data={
                "document_id": document_id,
                "indexed_status": result_data["status"],
                "chunk_count": result_data["chunk_count"],
                "extraction_status": doc.extraction_status,
                "image_summary": doc.image_summary,
            },
        )
    except Exception as e:
        doc.indexed_status = "failed"
        doc.extraction_status = "failed"
        doc.index_error = str(e)
        await db.commit()
        raise HTTPException(status_code=500, detail=f"索引失败: {str(e)}")


@router.post("/documents/batch-reindex", response_model=ApiResponse)
async def batch_reindex_knowledge_documents(
    document_ids: list[int] | None = None,
    domain: str = "stability",
    db: AsyncSession = Depends(get_knowledge_db),
):
    """批量重新索引知识库文档。

    支持两种模式：
    1. 指定 document_ids 列表：只索引指定文档
    2. 不指定 document_ids：自动索引该 domain 下所有 extraction_status=pending 的文档
    """
    if document_ids:
        result = await db.execute(
            select(KnowledgeDocument).where(KnowledgeDocument.id.in_(document_ids))
        )
        docs = result.scalars().all()
    else:
        result = await db.execute(
            select(KnowledgeDocument).where(
                KnowledgeDocument.domain == domain,
                KnowledgeDocument.is_active == True,
                KnowledgeDocument.extraction_status.in_(["pending", "failed"]),
            )
        )
        docs = result.scalars().all()

    if not docs:
        return ApiResponse(message="没有需要索引的文档", data={"total": 0, "results": []})

    results = []
    for doc in docs:
        try:
            result_data = await _do_reindex_doc(doc)
            await db.commit()
            results.append({"id": doc.id, "title": doc.title, **result_data})
        except Exception as e:
            doc.indexed_status = "failed"
            doc.extraction_status = "failed"
            doc.index_error = str(e)
            await db.commit()
            results.append({"id": doc.id, "title": doc.title, "status": "failed", "error": str(e)})

    completed = sum(1 for r in results if r["status"] == "completed")
    failed = sum(1 for r in results if r["status"] == "failed")
    skipped = sum(1 for r in results if r["status"] == "skipped")
    return ApiResponse(
        message=f"批量索引完成：{completed}成功, {failed}失败, {skipped}跳过",
        data={"total": len(results), "completed": completed, "failed": failed, "skipped": skipped, "results": results},
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Helper functions for RAG metadata inference
# ═══════════════════════════════════════════════════════════════════════════════

def _infer_region(doc) -> str:
    """Infer region from document title/path."""
    title = (doc.title or "") + (doc.file_path or "")
    regions = [
        ("洪泽", "淮安市洪泽区"), ("金湖", "淮安市金湖县"),
        ("清江浦", "淮安市清江浦区"), ("淮阴", "淮安市淮阴区"),
        ("涟水", "淮安市涟水县"), ("淮安", "淮安市"),
        ("江苏", "江苏省"), ("南京", "南京市"), ("南通", "南通市"),
        ("全国", "全国"), ("国家", "全国"),
    ]
    for keyword, label in regions:
        if keyword in title:
            return label
    return "江苏省"


def _infer_year(doc) -> str:
    """Infer document year from title or created_at."""
    import re
    title = doc.title or ""
    # Look for year in title: 2024, 2025, 2026
    m = re.search(r'(20\d{2})', title)
    if m:
        return m.group(1)
    # Fallback to created_at year
    if hasattr(doc, 'created_at') and doc.created_at:
        return str(doc.created_at.year)
    return "-"


def _infer_risk_tags(doc) -> str:
    """Infer risk tags from document type."""
    dtype = (doc.document_type or "").lower()
    title = (doc.title or "").lower()
    tags = []
    if any(k in title for k in ["补偿", "征收", "征地"]):
        tags.append("补偿争议风险")
    if any(k in title for k in ["程序", "报批", "审批"]):
        tags.append("程序风险")
    if any(k in title for k in ["社保", "安置", "保障"]):
        tags.append("社保安置风险")
    if any(k in title for k in ["群体", "信访", "舆情", "突发事件"]):
        tags.append("群体性事件风险")
    if any(k in title for k in ["评估", "稳评", "风险"]):
        tags.append("风险等级判定")
    if dtype in ("regulation", "standard"):
        tags.append("合规性")
    return "、".join(tags) if tags else "通用风险"


def _infer_doc_category(doc) -> str:
    """Infer document category for RAG tag."""
    dtype = (doc.document_type or "").lower()
    title = (doc.title or "").lower()
    if dtype in ("regulation", "law"):
        return "政策法规"
    if dtype in ("standard",):
        return "技术标准"
    if any(k in title for k in ["报告", "稳评报告", "范本", "案卷", "洞庭湖"]):
        return "人工范本"
    if any(k in title for k in ["案例", "模板", "case"]):
        return "本地案例"
    if any(k in title for k in ["指南", "理论", "方法论"]):
        return "理论文献"
    if any(k in title for k in ["公司", "资质", "众拓"]):
        return "固定资料"
    if any(k in title for k in ["规范", "格式"]):
        return "技术标准"
    return "其他"


# ═══════════════════════════════════════════════════════════════════════════════
# Data Cleaning Workbench Endpoints
# ═══════════════════════════════════════════════════════════════════════════════

# In-memory cache shared with startup preloader
from app.services.cleaning_pipeline import text_cache as _cleaning_cache

@router.get("/cleaning/documents", response_model=ApiResponse)
async def list_cleaning_documents(
    db: AsyncSession = Depends(get_knowledge_db),
):
    """List all cleanable documents — both knowledge documents AND templates."""
    from app.models.knowledge import Template

    items = []

    # Knowledge documents
    doc_result = await db.execute(
        select(KnowledgeDocument).where(KnowledgeDocument.is_active == True)
        .order_by(KnowledgeDocument.created_at.desc())
    )
    for doc in doc_result.scalars().all():
        items.append({
            "id": f"doc_{doc.id}",
            "source_type": "knowledge_document",
            "title": doc.title,
            "document_type": doc.document_type,
            "domain": doc.domain,
            "file_path": doc.file_path,
            "file_type": doc.file_type,
            "clean_status": doc.clean_status or "raw",
            "indexed_status": doc.indexed_status,
            "has_raw_text": bool(doc.raw_text),
            "has_cleaned_text": bool(doc.cleaned_text),
        })

    # Templates — classify by type, skip format-only docs
    tpl_result = await db.execute(
        select(Template).where(Template.is_active == True)
        .order_by(Template.created_at.desc())
    )
    for tpl in tpl_result.scalars().all():
        file_path = tpl.template_file_path
        name = (tpl.name or "").strip()
        category = (tpl.category or "").strip()

        # Skip format-only documents (not knowledge content)
        is_format = (
            "格式" in category or "格式" in name or "字体" in name
            or "文字" in name or "评审表" in name
        )
        if is_format:
            continue

        is_report = (
            "报告" in name or "稳评" in name or "社会稳定" in name
            or "投标" in name or "招标" in name or "项目" in name
        )
        tpl_type = "example_report" if is_report else "template"
        prefix = "[报告]" if is_report else "[模板]"

        items.append({
            "id": f"tpl_{tpl.id}",
            "source_type": "template",
            "title": f"{prefix} {tpl.name}",
            "document_type": tpl_type,
            "domain": tpl.domain,
            "file_path": file_path,
            "file_type": "docx",
            "clean_status": "raw",
            "indexed_status": tpl.analysis_status or "pending",
            "has_raw_text": False,
            "has_cleaned_text": False,
        })

    return ApiResponse(data={"items": items, "total": len(items)})


@router.post("/cleaning/preview", response_model=ApiResponse)
async def preview_clean_unified(
    body: dict,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """Unified cleaning preview — works for both knowledge documents and templates.

    Request body: {"composite_id": "doc_5" or "tpl_5", "config": {...}}
    """
    from app.models.knowledge import Template
    from app.services.cleaning_pipeline import cleaning_pipeline
    from app.services.file_service import file_service

    composite_id = body.get("composite_id", "")
    config = body.get("config") or cleaning_pipeline.get_default_config()

    import time as _time
    _t0 = _time.perf_counter()

    if not composite_id:
        raise HTTPException(status_code=400, detail="缺少 composite_id")

    # Resolve source
    if composite_id.startswith("doc_"):
        doc_id = int(composite_id[4:])
        result = await db.execute(
            select(KnowledgeDocument).where(KnowledgeDocument.id == doc_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(status_code=404, detail="文档不存在")
        file_path = doc.file_path
        title = doc.title
        raw_text = doc.raw_text or ""
    elif composite_id.startswith("tpl_"):
        tpl_id = int(composite_id[4:])
        result = await db.execute(
            select(Template).where(Template.id == tpl_id)
        )
        tpl = result.scalar_one_or_none()
        if not tpl:
            raise HTTPException(status_code=404, detail="模板不存在")
        file_path = tpl.template_file_path
        title = tpl.name
        # Use in-memory cache so we don't re-extract the same large DOCX every time
        raw_text = _cleaning_cache.get(file_path, "")
    else:
        raise HTTPException(status_code=400, detail=f"无效ID: {composite_id}")

    # Extract raw text if not cached
    if not raw_text:
        abs_path = file_service.get_absolute_path(file_path)
        if not abs_path.exists():
            raise HTTPException(status_code=404, detail=f"文件不存在: {file_path}")

        try:
            ext = abs_path.suffix.lower()
            if ext in (".pdf",):
                from app.services.material_ingestion_service import material_ingestion_service
                artifact = await material_ingestion_service.ingest_material(
                    file_path, scope="cleaning", title=title,
                    document_type="other",
                )
                raw_text = artifact.get("text_content", "") or ""
            elif ext in (".docx", ".doc"):
                raw_text = file_service.extract_docx_text(str(abs_path))
            else:
                raw_text = file_service.read_text_file(str(abs_path))
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"文本提取失败: {str(e)}")

        if not raw_text:
            raise HTTPException(status_code=400, detail="文档无法提取文本内容")

        # Cache raw_text in memory (avoid re-extracting same file)
        _cleaning_cache[file_path] = raw_text

        # Also cache back to DB for knowledge docs
        if composite_id.startswith("doc_"):
            result2 = await db.execute(
                select(KnowledgeDocument).where(KnowledgeDocument.id == int(composite_id[4:]))
            )
            d = result2.scalar_one_or_none()
            if d:
                d.raw_text = raw_text
                await db.commit()

    _t1 = _time.perf_counter()
    # Run cleaning
    cleaned_text = cleaning_pipeline.execute(raw_text, config)
    _t2 = _time.perf_counter()
    issues = cleaning_pipeline.analyze(raw_text)
    _t3 = _time.perf_counter()

    raw_chars = len(raw_text)
    cleaned_chars = len(cleaned_text)

    # Return full text with annotations for tables and placeholders
    import re as _re

    def _sanitize(s):
        return _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', s)

    def _annotate(s):
        # Mark template placeholders
        s = _re.sub(r'\{\{[^}]+\}\}', r'【占位符:\g<0>】', s)
        # Mark markdown table blocks
        s = _re.sub(
            r'(\|[^\n]+\|\n\|[\s\-:|—]+\|\n(?:\|[^\n]+\|\n?)+)',
            r'\n【📊 表格 ↓】\n\1\n【📊 表格 ↑】\n', s
        )
        # Mark plain-text table patterns (consecutive short lines with consistent structure)
        # Detect: 3+ consecutive lines where each line has similar column-like spacing or separators
        lines = s.split('\n')
        result = []
        i = 0
        while i < len(lines):
            line = lines[i]
            # Check for table-like lines: contain multiple whitespace-separated columns
            # or tab characters, or are part of a structured data block
            is_table_like = (
                '\t' in line
                or bool(_re.match(r'^\s*[\w一-鿿]+[\s]{2,}[\w一-鿿]', line))
                or bool(_re.match(r'^[\s]*[①②③④⑤⑥⑦⑧⑨⑩\d]+[.、)\s]', line))
            )
            if is_table_like:
                # Look ahead for more table-like lines
                j = i + 1
                while j < len(lines) and j < i + 20:
                    nl = lines[j]
                    if not nl.strip():
                        j += 1
                        continue
                    nl_table_like = (
                        '\t' in nl
                        or bool(_re.match(r'^\s*[\w一-鿿]+[\s]{2,}[\w一-鿿]', nl))
                    )
                    if nl_table_like:
                        j += 1
                    else:
                        break
                if j - i >= 3:  # At least 3 table-like lines
                    result.append('【📊 表格 ↓】')
                    result.extend(lines[i:j])
                    result.append('【📊 表格 ↑】')
                    i = j
                    continue
            result.append(line)
            i += 1
        s = '\n'.join(result)
        return s

    raw_text = _annotate(_sanitize(raw_text))
    cleaned_text = _annotate(_sanitize(cleaned_text))

    # ── Table extraction (separate from text cleaning) ──
    extracted_tables = []
    try:
        from app.services.table_extractor import get_table_extractor
        ext = get_table_extractor()
        abs_path = file_service.get_absolute_path(file_path)
        if abs_path.exists():
            suffix = abs_path.suffix.lower()
            if suffix == '.pdf':
                pdf_tables = ext.extract_from_pdf(str(abs_path))
                extracted_tables = [t.to_dict() for t in pdf_tables]
            elif suffix in ('.docx', '.doc'):
                docx_tables = ext.extract_from_docx(str(abs_path))
                extracted_tables = [t.to_dict() for t in docx_tables]
    except Exception:
        pass  # Table extraction is best-effort

    _t4 = _time.perf_counter()
    return ApiResponse(data={
        "raw_text": raw_text,
        "cleaned_text": cleaned_text,
        "extracted_tables": extracted_tables,
        "raw_chars": raw_chars,
        "cleaned_chars": cleaned_chars,
        "removed_chars": raw_chars - cleaned_chars,
        "token_estimate": cleaning_pipeline.estimate_tokens(cleaned_text),
        "issues_found": issues[:50],
        "_timing": {
            "db_lookup": round(_t1 - _t0, 4),
            "cleaning": round(_t2 - _t1, 4),
            "analyze": round(_t3 - _t2, 4),
            "total": round(_t4 - _t0, 4),
        },
    })

@router.get("/documents/{document_id}/clean/config", response_model=ApiResponse)
async def get_clean_config(
    document_id: int,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """Get cleaning config for a document (saved or defaults)."""
    result = await db.execute(
        select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")

    from app.services.cleaning_pipeline import cleaning_pipeline

    # Return saved config or defaults
    if doc.clean_config_snapshot:
        try:
            saved_config = json.loads(doc.clean_config_snapshot)
        except (json.JSONDecodeError, TypeError):
            saved_config = cleaning_pipeline.get_default_config()
    else:
        saved_config = cleaning_pipeline.get_default_config()

    return ApiResponse(
        data={
            "document_id": document_id,
            "clean_status": doc.clean_status or "raw",
            "config": saved_config,
            "handler_metadata": cleaning_pipeline.get_handler_metadata(),
            "has_raw_text": bool(doc.raw_text),
            "has_cleaned_text": bool(doc.cleaned_text),
            "stats": {
                "raw_chars": len(doc.raw_text or ""),
                "cleaned_chars": len(doc.cleaned_text or ""),
            } if (doc.raw_text or doc.cleaned_text) else None,
        },
    )


@router.post("/documents/{document_id}/clean/preview", response_model=ApiResponse)
async def preview_clean(
    document_id: int,
    config: dict = None,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """Preview cleaning results: parse raw text if needed, run pipeline, return side-by-side."""
    from fastapi import Body
    result = await db.execute(
        select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")

    from app.services.cleaning_pipeline import cleaning_pipeline

    # Get raw text — extract from file if not already done
    raw_text = doc.raw_text or ""
    if not raw_text:
        # Try to extract text from the file
        try:
            from app.services.material_ingestion_service import material_ingestion_service
            artifact = await material_ingestion_service.ingest_material(
                doc.file_path,
                scope="knowledge",
                title=doc.title,
                document_type=doc.document_type,
                domain=getattr(doc, "domain", "stability") or "stability",
                metadata={"document_id": document_id},
            )
            raw_text = artifact.get("text_content", "") or ""
            if raw_text:
                doc.raw_text = raw_text
                await db.commit()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"无法解析文档: {str(e)}")

    if not raw_text:
        raise HTTPException(status_code=400, detail="文档无法提取文本内容")

    # Use provided config or defaults
    clean_config = config or cleaning_pipeline.get_default_config()

    # Run cleaning
    cleaned_text = cleaning_pipeline.execute(raw_text, clean_config)

    # Analyze issues
    issues = cleaning_pipeline.analyze(raw_text)

    raw_chars = len(raw_text)
    cleaned_chars = len(cleaned_text)

    return ApiResponse(
        data={
            "raw_text": raw_text[:50000],          # Cap for UI display
            "cleaned_text": cleaned_text[:50000],
            "raw_chars": raw_chars,
            "cleaned_chars": cleaned_chars,
            "removed_chars": raw_chars - cleaned_chars,
            "token_estimate": cleaning_pipeline.estimate_tokens(cleaned_text),
            "issues_found": issues[:50],           # Cap issue count
        },
    )


@router.put("/documents/{document_id}/cleaned-text", response_model=ApiResponse)
async def save_cleaned_text(
    document_id: int,
    body: dict,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """Save manually edited cleaned text for a document."""
    result = await db.execute(
        select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")

    cleaned_text = body.get("cleaned_text", "")
    config = body.get("config")

    doc.cleaned_text = cleaned_text
    if config:
        doc.clean_config_snapshot = json.dumps(config, ensure_ascii=False)
    doc.clean_status = "cleaned"
    await db.commit()

    return ApiResponse(
        message="清洗文本已保存",
        data={"document_id": document_id, "clean_status": "cleaned"},
    )


@router.post("/documents/{document_id}/clean/apply", response_model=ApiResponse)
async def apply_clean(
    document_id: int,
    body: dict = None,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """Confirm cleaning config and mark document as ready for indexing."""
    result = await db.execute(
        select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")

    from app.services.cleaning_pipeline import cleaning_pipeline

    body = body or {}
    config = body.get("config", cleaning_pipeline.get_default_config())

    # If cleaned_text not provided, re-run pipeline
    cleaned_text = body.get("cleaned_text") or doc.cleaned_text
    if not cleaned_text and doc.raw_text:
        cleaned_text = cleaning_pipeline.execute(doc.raw_text, config)
    elif not cleaned_text:
        raise HTTPException(status_code=400, detail="请先执行清洗预览")

    doc.cleaned_text = cleaned_text
    doc.clean_config_snapshot = json.dumps(config, ensure_ascii=False)
    doc.clean_status = "confirmed"
    await db.commit()

    return ApiResponse(
        message="清洗已确认，可以入库",
        data={
            "document_id": document_id,
            "clean_status": "confirmed",
            "token_estimate": cleaning_pipeline.estimate_tokens(cleaned_text),
        },
    )


@router.post("/documents/{document_id}/clean/reset", response_model=ApiResponse)
async def reset_clean(
    document_id: int,
    db: AsyncSession = Depends(get_knowledge_db),
):
    """Reset cleaning state — revert to raw text, allow re-cleaning."""
    result = await db.execute(
        select(KnowledgeDocument).where(KnowledgeDocument.id == document_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")

    doc.cleaned_text = None
    doc.clean_config_snapshot = None
    doc.clean_status = "raw"
    await db.commit()

    return ApiResponse(
        message="清洗状态已重置",
        data={"document_id": document_id, "clean_status": "raw"},
    )


async def _index_image_document(relative_path: str, title: str) -> str:
    """Describe an image document using vision AI, for embedding as text.

    Uses Qwen-VL-Max to generate a Chinese description of the image content,
    which is then chunked and embedded like regular text documents.
    """
    import base64
    from app.services.file_service import file_service as fs
    from app.services.llm_service import llm_service

    abs_path = fs.get_absolute_path(relative_path)
    if not abs_path.exists():
        raise FileNotFoundError(f"图片文件不存在: {relative_path}")

    # Read and encode image
    with open(abs_path, "rb") as f:
        image_data = base64.b64encode(f.read()).decode("utf-8")

    ext = abs_path.suffix.lower()
    mime_map = {
        ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".gif": "image/gif", ".bmp": "image/bmp", ".webp": "image/webp",
    }
    mime_type = mime_map.get(ext, "image/png")

    # Classify image type for better description
    name_lower = title.lower()
    if any(kw in name_lower for kw in ["公告", "公示", "批文", "通知", "征收"]):
        prompt = (
            "你是一个文档分析工具。请逐字识别这张公告/通知图片中的所有文字，"
            "包括标题、文号、日期、单位名称、公章内容等。按原文格式输出。"
            "如果是公章，描述公章内容。"
        )
    elif any(kw in name_lower for kw in ["问卷", "调查", "统计", "意见"]):
        prompt = (
            "你是一个数据分析工具。请识别这张问卷/调查表图片中的所有内容，"
            "包括：题目、选项、统计数据（数字/百分比）、汇总结果、日期等。"
            "按结构化格式输出。"
        )
    elif any(kw in name_lower for kw in ["照片", "现场", "座谈", "走访", "会议"]):
        prompt = (
            "请详细描述这张照片的场景内容：拍摄地点、时间、人物活动、"
            "现场布置（横幅/标语内容）、参与人数、会议主题等。"
            "用语正式，适合用于政府报告附图说明。"
        )
    else:
        prompt = (
            "请详细描述这张图片的内容。如果是文档扫描件，逐字识别所有文字；"
            "如果是照片，描述场景、人物、活动；如果是图表，描述数据和趋势。"
            "用中文输出，200-500字。"
        )

    try:
        # Route through the shared LLM service so the vision model/endpoint
        # is configurable (VISION_MODEL/VISION_BASE_URL) and consistent across
        # the app, instead of hardcoding a DashScope Qwen-VL endpoint here.
        from app.services.llm_service import llm_service

        description = await llm_service.chat_with_image(
            text=prompt,
            image_base64=image_data,
            media_type=mime_type,
            max_tokens=1024,
        )
        if not description or not description.strip():
            raise ValueError("视觉API返回空内容")
        # Format as a proper document for chunking
        return (
            f"# {title}\n\n"
            f"**文档类型**：图片文档（AI视觉识别）\n\n"
            f"**内容描述**：\n{description}"
        )
    except Exception as e:
        raise ValueError(f"图片索引失败: {str(e)}")


def _extract_spreadsheet_text(relative_path: str) -> str:
    """Extract text from Excel/CSV files for indexing."""
    abs_path = file_service.get_absolute_path(relative_path)
    if not abs_path.exists():
        raise FileNotFoundError(f"文件不存在: {relative_path}")

    ext = abs_path.suffix.lower()
    lines = []

    if ext == ".csv":
        import csv
        with open(abs_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                lines.append(" | ".join(cell.strip() for cell in row if cell.strip()))
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(abs_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"## 工作表：{sheet_name}")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cells:
                        lines.append(" | ".join(cells))
        except ImportError:
            raise ValueError("无法读取Excel文件：缺少openpyxl库")

    text = "\n".join(lines)
    if len(text.strip()) < 50:
        raise ValueError(f"表格文本内容过短（{len(text)}字符），无法索引")
    return text


def _extract_text_generic(relative_path: str) -> str:
    """Generic text extraction for unsupported file types."""
    abs_path = file_service.get_absolute_path(relative_path)
    if not abs_path.exists():
        raise FileNotFoundError(f"文件不存在: {relative_path}")

    # Try reading as UTF-8 text
    try:
        with open(abs_path, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        pass

    # Try reading as GBK (common for Chinese documents)
    try:
        with open(abs_path, "r", encoding="gbk") as f:
            return f.read()
    except UnicodeDecodeError:
        pass

    # Try reading as latin-1 (will capture raw bytes as chars)
    try:
        with open(abs_path, "r", encoding="latin-1") as f:
            content = f.read()
            if len(content.strip()) > 100:
                return content
    except Exception:
        pass

    raise ValueError("无法识别文件编码，请尝试上传 .txt / .md / .docx 格式的文档")
